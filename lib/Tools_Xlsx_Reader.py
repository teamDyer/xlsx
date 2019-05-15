"""
"""

from pprint import pprint
from collections import OrderedDict
import openpyxl

sample = "sample.xlsx"

def example():
    """
    """
    ##nested dictionary all of the data comes into
    data = OrderedDict()
    ##load the file 
    workbook = openpyxl.load_workbook(filename=sample, read_only=True)
    ##grab all of the runs
    runs = workbook._sheets
    ##assume the structure based on the first run's header
    structure = []
    run = runs[0]
    for column in range(1,run._max_column+1):
        cell = run.cell(row=1,column=column)
        structure.append(cell.value)
    ##add all of the data from each run into the nested dictionary
    for run in runs:
        print(run)
        sheet = OrderedDict()
        ##add each test individually
        for row in range(2,run._max_row+1):
            ##the test number is stored in the first column of each row
            testnumber = run.cell(row=row,column=1).value
            test = OrderedDict()
            for column in range(1,run._max_column+1):
                attribute = structure[column-1]
                test[attribute] = run.cell(row=row,column=column).value
            sheet[testnumber] = test
        data[run.title] = sheet
    return data
